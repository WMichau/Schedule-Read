import sharepy
import openpyxl
import sys, getopt
import re
from sharepy import connect
from sharepy import SharePointSession
from pathlib import Path
from datetime import datetime

def Downlaod_File():
    root_url = "https://grupawppl.sharepoint.com"
    full_url = "https://grupawppl.sharepoint.com/sites/LoaSzydercw/Shared Documents/ZOP/Dyżury aktualne.xlsx"

    user='Michal.Wrycza@grupawp.pl'
    password='Wtfwtfwtf132@'

    s = sharepy.connect(root_url, user, password)
 
    r = s.getfile(full_url, \
        filename = "Dyżury aktualne.xlsx")

def Format_Date():
    xlsx_file = Path('./', 'Dyżury aktualne.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    for i in range(len(wb_obj.sheetnames)):
        sheet = wb_obj[wb_obj.sheetnames[i]]
        for cell in sheet['B']:
            if cell.value is not None:
                cell.value = str(cell.value).strip('00:00:00').replace('/', '-')
    wb_obj.save('Dyżury aktualne.xlsx')

def Read_File(name, sheet):
    xlsx_file = Path('./', 'Dyżury aktualne.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj[wb_obj.sheetnames[-int(sheet)]]

    morning = sheet["F2"].value
    afternoon = sheet["H2"].value
    night = sheet["J2"].value
    lotny = sheet["E2"].value
    day_name = "B"

    for row in sheet.iter_rows():
        for cell in row:
            if name in str(cell.value):
                if cell.column_letter == "F":
                    print(sheet.cell(row = cell.row, column = cell.column - 4).value, end=' ')
                    print(sheet.cell(row = cell.row, column = cell.column - 3).value, end=' ')
                    print(morning, end=' ')
                    if sheet.cell(row = cell.row, column = cell.column - 3).value == 'PON':
                        print("↓↓↓ " + sheet.cell(row = cell.row -2, column = cell.column + 4).value, end=' ')     
                    else:
                        print("↓↓↓ " + sheet.cell(row = cell.row -1, column = cell.column + 4).value, end=' ')                        
                    print("↑↑↑ " + sheet.cell(row = cell.row, column = cell.column + 2).value)
                elif cell.column_letter == "H":
                    print(sheet.cell(row = cell.row, column = cell.column - 6).value, end=' ')
                    print(sheet.cell(row = cell.row, column = cell.column - 5).value, end=' ')
                    print(afternoon, end=' ')
                    print("↓↓↓ " + sheet.cell(row = cell.row, column = cell.column - 2).value, end=' ')
                    print("↑↑↑ " + sheet.cell(row = cell.row, column = cell.column + 2).value)
                elif cell.column_letter == "J":
                    print(sheet.cell(row = cell.row, column = cell.column - 8).value, end=' ')
                    print(sheet.cell(row = cell.row, column = cell.column - 7).value, end=' ')
                    print(night, end=' ')
                    print("↓↓↓ " + sheet.cell(row = cell.row, column = cell.column - 2).value, end=' ')
                    if sheet.cell(row = cell.row + 2, column = cell.column).value == None:
                        print(' ')
                        break
                    else:
                        if (sheet.cell(row = cell.row, column = cell.column - 7).value == 'NIE'):
                            print("↑↑↑ " + sheet.cell(row = cell.row + 2, column = cell.column - 4).value)
                        else:
                            print("↑↑↑ " + sheet.cell(row = cell.row + 1, column = cell.column - 4).value)
                elif cell.column_letter == "E":
                    print(sheet.cell(row = cell.row, column = cell.column - 3).value, end=' ')
                    print(sheet.cell(row = cell.row, column = cell.column - 2).value, end=' ')
                    print(lotny)

# def Args(argv):
#     name = ''
    
#     try:
#         opts, args = getopt.getopt(argv, "n")
#     except:
#         print('Grafik.py without arguments for all or -n name')
#         sys.exit(2)
#     for opt, arg in opts:
#         if opt == '-n':
#             name = arg
#         else:
#             name = 'dupa'
#     print(name)

if __name__ == '__main__': 
    Downlaod_File()
    Format_Date()
    
    name = sys.argv[1]
    sheet = sys.argv[2]

    Read_File(name, sheet)
    


